#!/usr/bin/env python3
from __future__ import annotations

import argparse
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import AreaChart, DoughnutChart, LineChart, PieChart, Reference, ScatterChart, Series
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils.datetime import CALENDAR_MAC_1904
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.protection import WorkbookProtection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image as PILImage


def prepare(path: Path) -> None:
    wb = Workbook()
    wb.epoch = CALENDAR_MAC_1904
    wb.properties.title = "Fixture 1.5"
    wb.properties.subject = "Interopérabilité"
    wb.properties.creator = "openpyxl"
    wb.properties.description = "Classeur source pour lua-xlsx"
    wb.properties.keywords = "lua, xlsx, interop"
    wb.properties.category = "Tests"
    wb.security = WorkbookProtection(workbookPassword="secret", lockStructure=True)

    ws = wb.active
    ws.title = "Entrée"
    ws.append(["Nom", "Actif", "Date", "Formule"])
    ws.append(["Zoé", False, datetime(2025, 1, 2, 3, 4, 5), "=SUM(1,2)"])
    ws.append(["Autre", True, datetime(2025, 1, 3, 4, 5, 6), 7])
    ws["A6"] = CellRichText(
        TextBlock(InlineFont(b=True, color="FFFF0000"), "Attention : "),
        TextBlock(InlineFont(i=True), "texte enrichi openpyxl"),
    )
    ws["A2"].hyperlink = "https://example.org/source?a=1&b=2"
    ws["A2"].hyperlink.tooltip = "Source externe"
    ws["B2"].font = Font(
        name="DejaVu Sans", size=13, bold=True, italic=True,
        underline="single", strike=True, color="FF123456",
    )
    ws["B2"].fill = PatternFill(fill_type="solid", fgColor="FFABCDEF")
    ws["B2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["B2"].border = Border(
        left=Side(style="thin", color="FFFF0000"),
        right=Side(style="double"),
        top=Side(style="dashed", color="FF00AA00"),
        bottom=Side(style="dotted"),
    )
    ws["B2"].protection = Protection(locked=False)
    ws["D2"].protection = Protection(hidden=True)
    ws.protection.sheet = True
    ws.protection.password = "feuille"
    ws.protection.selectUnlockedCells = False
    ws.protection.autoFilter = False

    ws.column_dimensions["A"].width = 20
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = "A1:D3"
    ws.merge_cells("A4:C4")
    ws["A4"] = "Fusion openpyxl"
    ws.sheet_properties.tabColor = "5B9BD5"
    ws["C2"].comment = Comment("Commentaire openpyxl", "Interop")
    ws.row_dimensions[5].hidden = True
    ws.column_dimensions["E"].hidden = True

    dv_list = DataValidation(
        type="list", formula1='"Oui,Non,En attente"', allow_blank=True,
        showInputMessage=True, showErrorMessage=True, promptTitle="Choix",
        prompt="Sélectionner", errorTitle="Erreur", error="Valeur invalide",
    )
    ws.add_data_validation(dv_list)
    dv_list.add("A2:A10")
    dv_whole = DataValidation(type="whole", operator="between", formula1="0", formula2="100")
    ws.add_data_validation(dv_whole)
    dv_whole.add("B2:B10")

    ws.conditional_formatting.add(
        "B2:B10",
        CellIsRule(
            operator="lessThan", formula=["0"],
            font=Font(color="FF9C0006"),
            fill=PatternFill(fill_type="solid", fgColor="FFFFC7CE"),
        ),
    )
    ws.conditional_formatting.add(
        "B2:B10",
        ColorScaleRule(
            start_type="min", start_color="FFF8696B",
            mid_type="percentile", mid_value=50, mid_color="FFFFEB84",
            end_type="max", end_color="FF63BE7B",
        ),
    )
    ws.conditional_formatting.add(
        "C2:C10",
        DataBarRule(start_type="min", end_type="max", color="FF5B9BD5", showValue=True),
    )
    ws.conditional_formatting.add(
        "D2:D10",
        IconSetRule(icon_style="3TrafficLights1", type="num", values=[-1, 0, 3], showValue=True),
    )

    image_path = path.with_suffix(".png")
    PILImage.new("RGB", (20, 10), (30, 144, 255)).save(image_path)
    picture = XLImage(str(image_path))
    picture.width = 80
    picture.height = 40
    ws.add_image(picture, "F2")

    input_table = Table(displayName="InputTable", ref="A1:D3")
    input_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(input_table)

    line = LineChart()
    line.title = "Tendance openpyxl"
    line.add_data(Reference(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    line.set_categories(Reference(ws, min_col=1, min_row=2, max_row=3))
    ws.add_chart(line, "F8")

    pie = PieChart()
    pie.title = "Répartition openpyxl"
    pie.add_data(Reference(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=3))
    ws.add_chart(pie, "F24")

    doughnut = DoughnutChart()
    doughnut.title = "Anneau openpyxl"
    doughnut.holeSize = 65
    doughnut.add_data(Reference(ws, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    doughnut.set_categories(Reference(ws, min_col=1, min_row=2, max_row=3))
    ws.add_chart(doughnut, "F40")

    area = AreaChart()
    area.title = "Aire openpyxl"
    area.grouping = "stacked"
    area.add_data(Reference(ws, min_col=2, max_col=3, min_row=1, max_row=3), titles_from_data=True)
    area.set_categories(Reference(ws, min_col=1, min_row=2, max_row=3))
    ws.add_chart(area, "F56")

    scatter = ScatterChart()
    scatter.title = "Nuage openpyxl"
    scatter.series.append(Series(
        Reference(ws, min_col=2, min_row=2, max_row=3),
        Reference(ws, min_col=3, min_row=2, max_row=3),
        title="Mesures",
    ))
    ws.add_chart(scatter, "F72")

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.4
    ws.oddHeader.center.text = "OpenPyXL"
    ws.oddFooter.right.text = "Page &P / &N"
    ws.print_area = "A1:D10"
    ws.print_title_rows = "1:1"
    ws.print_title_cols = "A:A"
    ws.row_breaks.append(Break(id=4))
    ws.col_breaks.append(Break(id=3))

    target = wb.create_sheet("Cible")
    target["A1"] = "destination"
    hidden = wb.create_sheet("Masquée")
    hidden.sheet_state = "hidden"
    very_hidden = wb.create_sheet("Très masquée")
    very_hidden.sheet_state = "veryHidden"
    wb.active = 0
    wb.defined_names.add(DefinedName("ZoneNoms", attr_text="'Entrée'!$A$2:$A$10"))
    wb.defined_names.add(DefinedName("ValeurLocale", attr_text="'Entrée'!$B$2", localSheetId=0, hidden=True))
    wb.save(path)
    image_path.unlink(missing_ok=True)
    print(f"Fixture openpyxl créée : {path}")


def check(path: Path) -> None:
    wb = load_workbook(path, data_only=False, rich_text=True)
    assert wb.epoch == CALENDAR_MAC_1904
    assert wb.sheetnames == ["Résumé", "Cible", "Masquée", "Très masquée"]
    assert wb.properties.title == "Rapport 1.5"
    assert wb.properties.creator == "lua-xlsx"
    assert wb.properties.keywords == "lua, xlsx, interop"
    assert wb.security.lockStructure is True
    assert wb.security.workbookPassword is not None

    ws = wb["Résumé"]
    assert ws["A1"].value == "Rapport 1.5"
    assert str(next(iter(ws.merged_cells.ranges))) == "A1:D1"
    assert ws["A3"].value == "Élise & <test>"
    assert ws["B3"].value == 12.5
    assert ws["C3"].value == datetime(2024, 3, 15, 13, 45, 30)
    assert ws["D5"].value == "=SUM(B3:B4)"

    rich = ws["A7"].value
    assert isinstance(rich, CellRichText)
    assert str(rich) == "Attention : valeur à contrôler"
    assert rich[0].font.b is True and rich[0].font.color.rgb == "FFFF0000"
    assert rich[1].font.i is True

    assert ws.freeze_panes == "B3"
    assert ws.auto_filter.ref == "A2:D4"
    assert ws.column_dimensions["A"].width == 18.0
    assert ws.column_dimensions["B"].width == 14.0
    assert ws.row_dimensions[1].height == 28.0

    assert ws["A1"].font.bold is True
    assert ws["A1"].font.italic is True
    assert ws["A1"].font.name == "Liberation Sans"
    assert ws["A1"].font.sz == 14.0
    assert ws["A1"].font.underline == "double"
    assert ws["A1"].font.strike is True
    assert ws["A1"].font.color.rgb == "FF112233"
    assert ws["A1"].fill.fgColor.rgb == "FFD9EAF7"
    assert ws["A1"].alignment.horizontal == "center"
    assert ws["A1"].alignment.vertical == "center"
    assert ws["A1"].alignment.wrap_text is True
    assert ws["A1"].border.left.style == "thin"
    assert ws["A1"].border.left.color.rgb == "FFFF0000"
    assert ws["A1"].border.right.style == "double"
    assert ws["A1"].border.top.style == "dashed"
    assert ws["A1"].border.bottom.style == "dotted"
    assert ws["B3"].number_format == '#,##0.00 "€"'
    assert ws["C3"].font.bold is True
    assert ws["C3"].number_format == "yyyy-mm-dd hh:mm:ss"

    assert ws["B3"].protection.locked is False
    assert ws["D5"].protection.hidden is True
    assert ws.protection.sheet is True
    assert ws.protection.password is not None

    assert ws["D3"].hyperlink.target == "https://example.com/?a=1&b=2"
    assert ws["D3"].hyperlink.tooltip == "Ouvrir le site"
    assert ws["D4"].hyperlink.location == "'Cible'!A1"

    assert wb.active.title == "Résumé"
    assert ws.sheet_properties.tabColor.rgb == "FF4472C4"
    assert ws.row_dimensions[6].hidden is True
    assert ws.column_dimensions["E"].hidden is True
    assert ws["A4"].comment.text == "Valeur contrôlée par lua-xlsx"
    assert ws["A4"].comment.author == "Julien"

    validations = list(ws.data_validations.dataValidation)
    assert len(validations) == 2
    assert validations[0].type == "list" and str(validations[0].sqref) == "A3:A20"
    assert validations[0].formula1 == '"Oui,Non,En attente"'
    assert validations[1].type == "whole" and validations[1].operator == "between"

    all_rules = [rule for rules in ws.conditional_formatting._cf_rules.values() for rule in rules]
    kinds = {rule.type for rule in all_rules}
    assert {"cellIs", "colorScale", "dataBar", "iconSet", "top10", "aboveAverage"}.issubset(kinds)
    color_scale = next(rule for rule in all_rules if rule.type == "colorScale")
    assert color_scale.colorScale.color[1].rgb == "FFFFEB84"
    icon_set = next(rule for rule in all_rules if rule.type == "iconSet")
    assert icon_set.iconSet.iconSet == "3TrafficLights1"

    assert wb["Masquée"].sheet_state == "hidden"
    assert wb["Très masquée"].sheet_state == "veryHidden"
    names = {name.name: name for name in wb.defined_names.values()}
    assert names["ZoneNoms"].attr_text == "'Résumé'!$A$3:$A$20"
    local_name = ws.defined_names["ValeurLocale"]
    assert local_name.localSheetId == 0 and local_name.hidden is True

    assert len(ws._images) == 1
    assert ws._images[0].anchor._from.row == 7
    assert ws._images[0].anchor._from.col == 0
    assert list(ws.tables) == ["ResumeTable"]
    assert ws.tables["ResumeTable"].ref == "A2:D4"

    chart_names = {type(chart).__name__ for chart in ws._charts}
    assert chart_names == {"LineChart", "PieChart", "DoughnutChart", "AreaChart", "ScatterChart"}
    pie = next(chart for chart in ws._charts if type(chart).__name__ == "PieChart")
    assert pie.legend.position == "b"
    doughnut = next(chart for chart in ws._charts if type(chart).__name__ == "DoughnutChart")
    assert doughnut.holeSize == 65
    area = next(chart for chart in ws._charts if type(chart).__name__ == "AreaChart")
    assert area.grouping == "stacked"
    scatter = next(chart for chart in ws._charts if type(chart).__name__ == "ScatterChart")
    assert scatter.legend is None

    assert ws.page_setup.orientation == "landscape"
    assert ws.page_setup.paperSize == 9
    assert ws.page_setup.fitToWidth == 1
    assert ws.page_setup.fitToHeight == 0
    assert ws.page_margins.left == 0.4
    assert ws.oddHeader.center.text == "lua-xlsx 1.5"
    assert ws.oddFooter.right.text == "Page &P / &N"
    assert str(ws.print_area) == "'Résumé'!$A$1:$J$30"
    assert ws.print_title_rows == "$1:$2"
    assert ws.print_title_cols == "$A:$A"
    assert [brk.id for brk in ws.row_breaks.brk] == [4]
    assert [brk.id for brk in ws.col_breaks.brk] == [4]

    with zipfile.ZipFile(path) as archive:
        app = archive.read("docProps/app.xml").decode("utf-8")
        assert "lua-xlsx" in app and "Julien" in app

    data_wb = load_workbook(path, data_only=True)
    data_ws = data_wb["Résumé"]
    assert data_ws["C3"].value == datetime(2024, 3, 15, 13, 45, 30)
    assert data_ws["D5"].value == 19.75

    print("INTEROP OPENPYXL : PASS")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("action", choices=("prepare", "check"))
    parser.add_argument("path", type=Path)
    args = parser.parse_args()
    if args.action == "prepare":
        prepare(args.path)
    else:
        check(args.path)


if __name__ == "__main__":
    main()
